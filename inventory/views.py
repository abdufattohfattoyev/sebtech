from django.core.exceptions import ValidationError
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.http import require_GET
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import Sum, F, DecimalField, Q
from decimal import Decimal
from django_filters import rest_framework as filters

from .forms import PhoneForm, AccessoryForm, AccessoryAddQuantityForm, SupplierForm, SupplierPaymentForm
from .models import Phone, Accessory, AccessoryPurchaseHistory, ExternalSeller, DailySeller, PhoneModel, Supplier, \
    SupplierPaymentDetail, SupplierPayment
from shops.models import Shop

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.contrib.auth.decorators import login_required


def can_edit_inventory(user):
    """Faqat boss va finance tahrirlash huquqiga ega"""
    if not hasattr(user, 'userprofile'):
        return False
    return user.userprofile.role in ['boss', 'finance']


def get_redirect_url(request):
    """URL yo'naltirish manzilini aniqlash"""
    if "phone" in request.path:
        return "inventory:phone_list"
    elif "supplier" in request.path:
        return 'inventory:supplier_list'
    else:
        return "inventory:accessory_list"


def boss_or_finance_required(view_func):
    """Faqat boss yoki finance ruxsat beruvchi dekorator"""
    def wrapper(request, *args, **kwargs):
        if not can_edit_inventory(request.user):
            messages.error(
                request,
                "Sizda bu amalni bajarish huquqi yo'q. "
                "Faqat rahbar va moliyachi tahrirlashi mumkin."
            )
            return redirect(get_redirect_url(request))
        return view_func(request, *args, **kwargs)
    return wrapper


class PhoneFilter(filters.FilterSet):
    imei = filters.CharFilter(field_name='imei', lookup_expr='exact')
    status = filters.ChoiceFilter(field_name='status', choices=Phone.STATUS_CHOICES)

    class Meta:
        model = Phone
        fields = ['imei', 'status']


class AccessoryFilter(filters.FilterSet):
    code = filters.CharFilter(field_name='code', lookup_expr='exact')
    name = filters.CharFilter(field_name='name', lookup_expr='icontains')

    class Meta:
        model = Accessory
        fields = ['code', 'name']

@login_required
def dashboard(request):
    """Dashboard - faqat do'kondagi, ustadagi va qaytarilgan telefonlar"""
    shops = Shop.objects.prefetch_related('phones', 'accessories')
    status_filter = request.GET.get('status', '')

    if not shops.exists():
        messages.error(request, "Hozircha tizimda do'kon mavjud emas.")
        return redirect('shop:dashboard')

    total_phone_value = Decimal('0.00')
    total_accessory_value = Decimal('0.00')
    total_phone_count = 0
    total_accessory_count = 0
    total_phone_debt = Decimal('0.00')  # Faqat telefonlar qarzi
    total_our_money = Decimal('0.00')
    shop_stats = []

    # ✅ Boshlang'ich qarzni FAQAT BIR MARTA hisoblash (barcha taminotchilardan)
    all_suppliers_initial_debt = Supplier.objects.aggregate(
        total=Sum('initial_debt', output_field=DecimalField(max_digits=15, decimal_places=2))
    )['total'] or Decimal('0.00')

    for shop in shops:
        # ✅ Faqat ko'rsatish uchun filtrlangan telefonlar
        phones = shop.phones.filter(status__in=['shop', 'master', 'returned'])

        if status_filter:
            phones = phones.filter(status=status_filter)

        phone_count = phones.count()
        phone_cost_sum = phones.aggregate(
            total=Sum('cost_price', output_field=DecimalField(max_digits=15, decimal_places=2))
        )['total'] or Decimal('0.00')

        accessories = shop.accessories.all()
        accessory_count = accessories.aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        accessory_cost_sum = accessories.aggregate(
            total=Sum(F('purchase_price') * F('quantity'), output_field=DecimalField(max_digits=15, decimal_places=2))
        )['total'] or Decimal('0.00')

        # ✅ FAQAT shu shopdagi telefonlarning qarzi
        phone_debt = shop.phones.filter(
            source_type='supplier',
            payment_status__in=['debt', 'partial']
        ).aggregate(
            total=Sum('debt_balance', output_field=DecimalField(max_digits=15, decimal_places=2))
        )['total'] or Decimal('0.00')

        # ✅ Har bir do'konda FAQAT telefon qarzi ko'rsatiladi
        shop_supplier_debt = phone_debt

        # Hisob pulimiz = Telefonlar qiymati - Telefon qarzi
        our_money = phone_cost_sum - shop_supplier_debt

        total_phone_count += phone_count
        total_accessory_count += accessory_count
        total_phone_value += phone_cost_sum
        total_accessory_value += accessory_cost_sum
        total_phone_debt += phone_debt  # Faqat telefonlar qarzi
        total_our_money += our_money  # Har do'kon uchun hisoblanadi

        shop_stats.append({
            'shop': shop,
            'phone_count': phone_count,
            'accessory_count': accessory_count,
            'phone_cost_value': phone_cost_sum,
            'accessory_cost_value': accessory_cost_sum,
            'total_value': phone_cost_sum + accessory_cost_sum,
            'supplier_debt': shop_supplier_debt,  # Faqat telefonlar qarzi
            'our_money': our_money,
        })

    # ✅ UMUMIY QARZ = Barcha telefonlar qarzi + Boshlang'ich qarz (faqat 1 marta)
    total_supplier_debt = total_phone_debt + all_suppliers_initial_debt

    # ✅ UMUMIY HISOB PULIMIZ = Jami telefon qiymati - Umumiy qarz
    total_our_money = total_phone_value - total_supplier_debt

    all_phones = Phone.objects.filter(shop__in=shops)
    phones_in_shop = all_phones.filter(status='shop').count()
    phones_sold = all_phones.filter(status='sold').count()
    phones_in_repair = all_phones.filter(status='master').count()
    phones_returned = all_phones.filter(status='returned').count()
    phones_exchanged = all_phones.filter(status='exchanged_in').count()

    context = {
        'shops': shops,
        'shop_stats': shop_stats,
        'total_phones': total_phone_count,
        'total_accessories': total_accessory_count,
        'total_phone_value': total_phone_value,
        'total_accessory_value': total_accessory_value,
        'total_shops': shops.count(),
        'total_supplier_debt': total_supplier_debt,  # Umumiy: telefonlar + boshlang'ich
        'total_our_money': total_our_money,
        'status_filter': status_filter,
        'status_choices': Phone.STATUS_CHOICES,
        'phones_in_shop': phones_in_shop,
        'phones_sold': phones_sold,
        'phones_in_repair': phones_in_repair,
        'phones_returned': phones_returned,
        'phones_exchanged': phones_exchanged,
        'all_suppliers_initial_debt': all_suppliers_initial_debt,  # Debug uchun
    }
    return render(request, 'inventory/dashboard.html', context)


@login_required
def phone_list(request):
    """Telefonlar ro'yxati"""
    imei_query = request.GET.get('imei', '').strip()
    model_query = request.GET.get('model', '').strip()
    status_filter = request.GET.get('status', '').strip()
    shop_id = request.GET.get('shop_id', '').strip()
    page_number = request.GET.get('page', 1)

    phones = Phone.objects.all().select_related('shop', 'phone_model', 'memory_size').order_by('-id')

    if imei_query:
        phones = phones.filter(imei__icontains=imei_query)
    if model_query:
        phones = phones.filter(phone_model__id=model_query)
    if status_filter:
        phones = phones.filter(status=status_filter)
    if shop_id:
        phones = phones.filter(shop_id=shop_id)

    paginator = Paginator(phones, 12)
    page_obj = paginator.get_page(page_number)

    stats = {
        'total_phones': phones.count(),
        'displayed': len(page_obj),
        'phones_in_shop': phones.filter(status='shop').count(),
        'phones_sold': phones.filter(status='sold').count(),
    }

    phone_models = PhoneModel.objects.all().order_by('model_name')

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if is_ajax:
        phones_data = [{
            'id': phone.id,
            'phone_model': str(phone.phone_model),
            'memory_size': str(phone.memory_size),
            'shop_name': phone.shop.name,
            'imei': phone.imei,
            'condition_percentage': phone.condition_percentage,
            'cost_price': float(phone.cost_price),
            'sale_price': float(phone.sale_price) if phone.sale_price else None,
            'status': phone.status,
            'status_display': phone.get_status_display(),
            'image': phone.image.url if phone.image else None,
        } for phone in page_obj]

        response = JsonResponse({
            'phones': phones_data,
            'stats': stats,
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
        })
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response

    context = {
        'phones': page_obj,
        'phone_models': phone_models,
        'shops': Shop.objects.all(),
        'status_choices': Phone.STATUS_CHOICES,
        'imei_query': imei_query,
        'model_query': model_query,
        'status_filter': status_filter,
        'selected_shop': Shop.objects.filter(id=shop_id).first() if shop_id else None,
        'phones_in_shop': stats['phones_in_shop'],
        'phones_sold': stats['phones_sold'],
    }

    response = render(request, 'inventory/phone_list.html', context)
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@login_required
@boss_or_finance_required
def phone_create(request):
    """Yangi telefon qo'shish"""
    if request.method == 'POST':
        form = PhoneForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            try:
                phone = form.save(commit=False)
                phone.created_by = request.user
                phone.save()
                messages.success(request, "Telefon muvaffaqiyatli qo'shildi!")
                return redirect('inventory:phone_list')
            except Exception as e:
                messages.error(request, f"Telefon saqlashda xatolik: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, f"Xatolik: {error}")
                    else:
                        field_label = form[field].label or field
                        messages.error(request, f"{field_label}: {error}")
    else:
        form = PhoneForm(user=request.user)

    return render(request, 'inventory/phone_form.html', {
        'form': form,
        'is_edit': False,
    })


@login_required
@boss_or_finance_required
def phone_update(request, pk):
    """Telefonni tahrirlash"""
    phone = get_object_or_404(Phone, pk=pk)

    if request.method == 'POST':
        form = PhoneForm(request.POST, request.FILES, instance=phone, user=request.user)
        if form.is_valid():
            try:
                phone = form.save(commit=True)
                messages.success(request, "Telefon ma'lumotlari yangilandi!")
                return redirect('inventory:phone_list')
            except Exception as e:
                messages.error(request, f"Yangilashda xatolik: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, f"Xatolik: {error}")
                    else:
                        field_label = form[field].label or field
                        messages.error(request, f"{field_label}: {error}")
    else:
        form = PhoneForm(instance=phone, user=request.user)

    return render(request, 'inventory/phone_form.html', {
        'form': form,
        'phone': phone,
        'is_edit': True,
    })


@login_required
@boss_or_finance_required
def phone_delete(request, pk):
    """Telefonni o'chirish"""
    phone = get_object_or_404(Phone, pk=pk)

    if request.method == 'POST':
        try:
            phone.delete()
            messages.success(request, "Telefon muvaffaqiyatli o'chirildi!")
        except Exception as e:
            messages.error(request, f"O'chirishda xatolik: {str(e)}")
        return redirect('inventory:phone_list')

    return render(request, 'inventory/phone_confirm_delete.html', {'phone': phone})

@login_required
def accessory_list(request):
    """Aksessuarlar ro'yxati"""
    shops = Shop.objects.all()
    shop_id = request.GET.get('shop_id', '')
    code_query = request.GET.get('code', '').strip()
    name_query = request.GET.get('name', '').strip()
    stock_filter = request.GET.get('stock_level', '')

    accessories = Accessory.objects.select_related('supplier', 'created_by', 'shop')

    if shop_id:
        try:
            selected_shop = get_object_or_404(Shop, id=shop_id)
            accessories = accessories.filter(shop=selected_shop)
        except Shop.DoesNotExist:
            selected_shop = None
    else:
        selected_shop = None

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return accessory_search(request)

    if code_query:
        accessories = accessories.filter(code__icontains=code_query)
    if name_query:
        accessories = accessories.filter(name__icontains=name_query)

    stock_filters = {
        'low': Q(quantity__lt=5),
        'medium': Q(quantity__gte=5, quantity__lte=10),
        'high': Q(quantity__gt=10)
    }
    if stock_filter in stock_filters:
        accessories = accessories.filter(stock_filters[stock_filter])

    accessories = accessories.order_by('-created_at')

    total_accessory_cost = accessories.aggregate(
        total=Sum(F('purchase_price') * F('quantity'), output_field=DecimalField(max_digits=15, decimal_places=2))
    )['total'] or Decimal('0.00')

    for accessory in accessories:
        accessory.total_cost = accessory.purchase_price * accessory.quantity

    paginator = Paginator(accessories, 12)
    page_number = request.GET.get('page', 1)
    accessories_paginated = paginator.get_page(page_number)

    context = {
        'accessories': accessories_paginated,
        'shops': shops,
        'selected_shop': selected_shop,
        'code_query': code_query,
        'name_query': name_query,
        'stock_filter': stock_filter,
        'total_accessory_cost': total_accessory_cost,
    }
    return render(request, 'inventory/accessory_list.html', context)


@login_required
@boss_or_finance_required
def accessory_create(request):
    """Yangi aksessuar qo'shish"""
    if request.method == 'POST':
        form = AccessoryForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            try:
                accessory = form.save(commit=True)

                if hasattr(form, '_existing_accessory'):
                    quantity = form.cleaned_data.get('quantity', 0)
                    messages.success(request, f"✅ Mavjud aksessuar ustiga {quantity} dona qo'shildi!")
                else:
                    messages.success(request, f"✅ Yangi aksessuar '{accessory.name}' muvaffaqiyatli qo'shildi!")

                return redirect('inventory:accessory_list')

            except Exception as e:
                messages.error(request, f"❌ Aksessuar saqlashda xatolik: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, f"❌ {error}")
                    else:
                        field_label = form.fields.get(field, field).label or field
                        messages.error(request, f"❌ {field_label}: {error}")
    else:
        form = AccessoryForm(user=request.user)

    return render(request, 'inventory/accessory_form.html', {
        'form': form,
        'is_edit': False,
    })


@login_required
@boss_or_finance_required
def accessory_update(request, pk):
    """Aksessuarni tahrirlash"""
    accessory = get_object_or_404(Accessory, pk=pk)

    if request.method == 'POST':
        form = AccessoryForm(request.POST, request.FILES, instance=accessory, user=request.user)
        if form.is_valid():
            try:
                form.save(commit=True)
                messages.success(request, f"✅ '{accessory.name}' ma'lumotlari yangilandi!")
                return redirect('inventory:accessory_list')

            except Exception as e:
                messages.error(request, f"❌ Yangilashda xatolik: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, f"❌ {error}")
                    else:
                        field_label = form.fields.get(field, field).label or field
                        messages.error(request, f"❌ {field_label}: {error}")
    else:
        form = AccessoryForm(instance=accessory, user=request.user)

    return render(request, 'inventory/accessory_form.html', {
        'form': form,
        'accessory': accessory,
        'is_edit': True,
    })


@login_required
@boss_or_finance_required
def accessory_delete(request, pk):
    """Aksessuarni o'chirish"""
    accessory = get_object_or_404(Accessory, pk=pk)

    if request.method == 'POST':
        try:
            name = accessory.name
            accessory.delete()
            messages.success(request, f"✅ '{name}' muvaffaqiyatli o'chirildi!")
        except Exception as e:
            messages.error(request, f"❌ O'chirishda xatolik: {str(e)}")
        return redirect('inventory:accessory_list')

    return render(request, 'inventory/accessory_confirm_delete.html', {'accessory': accessory})


@login_required
@boss_or_finance_required
def accessory_add_quantity(request, pk):
    """Mavjud aksessuar ustiga soni qo'shish"""
    accessory = get_object_or_404(Accessory, pk=pk)

    if request.method == 'POST':
        form = AccessoryAddQuantityForm(request.POST)
        if form.is_valid():
            try:
                quantity = form.cleaned_data['quantity']
                accessory = form.save(accessory=accessory, user=request.user, commit=True)
                messages.success(request, f"✅ '{accessory.name}' ga {quantity} dona qo'shildi!")
                return redirect('inventory:accessory_list')

            except Exception as e:
                messages.error(request, f"❌ Xatolik: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, f"❌ {error}")
                    else:
                        field_label = form.fields.get(field, field).label or field
                        messages.error(request, f"❌ {field_label}: {error}")
    else:
        form = AccessoryAddQuantityForm()

    return render(request, 'inventory/accessory_add_quantity.html', {
        'accessory': accessory,
        'form': form
    })


@login_required
def accessory_quick_add(request):
    """Kod orqali tez qo'shish"""
    if request.method == 'POST':
        shop_id = request.POST.get('shop')
        code = request.POST.get('code', '').strip().zfill(4)
        quantity = request.POST.get('quantity', 0)
        purchase_price = request.POST.get('purchase_price')

        try:
            shop = Shop.objects.get(id=shop_id, owner=request.user)
            accessory = Accessory.objects.get(shop=shop, code=code)
            quantity = int(quantity)
            purchase_price = Decimal(purchase_price) if purchase_price else None

            if not purchase_price:
                return JsonResponse({'success': False, 'message': "Tannarx kiritilishi kerak!"})

            if quantity <= 0:
                return JsonResponse({'success': False, 'message': "Soni 0 dan katta bo'lishi kerak!"})

            AccessoryPurchaseHistory.objects.create(
                accessory=accessory,
                quantity=quantity,
                purchase_price=purchase_price,
                created_by=request.user
            )
            accessory.save()

            return JsonResponse({
                'success': True,
                'message': f"{accessory.name} ga {quantity} dona qo'shildi!",
                'new_quantity': accessory.quantity
            })

        except Shop.DoesNotExist:
            return JsonResponse({'success': False, 'message': "Do'kon topilmadi!"})
        except Accessory.DoesNotExist:
            return JsonResponse({'success': False, 'message': f"Kod {code} topilmadi!"})
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'message': "Noto'g'ri ma'lumot kiritildi!"})

    shops = Shop.objects.filter(owner=request.user)
    return render(request, 'inventory/accessory_quick_add.html', {'shops': shops})


@login_required
def accessory_detail(request, pk):
    """Aksessuar tafsilotlari"""
    accessory = get_object_or_404(Accessory, pk=pk, shop__owner=request.user)
    return render(request, 'inventory/accessory_detail.html', {'accessory': accessory})


@login_required
def shop_detail(request, shop_id):
    """Do'kon tafsilotlari"""
    shop = get_object_or_404(Shop, id=shop_id)

    phones = Phone.objects.filter(
        shop=shop,
        status__in=['shop', 'master', 'returned']
    ).select_related(
        'phone_model', 'memory_size', 'supplier', 'external_seller', 'created_by'
    ).order_by('-created_at')

    accessories = Accessory.objects.filter(shop=shop).select_related(
        'supplier', 'created_by'
    ).order_by('-created_at')

    phone_cost_value = phones.aggregate(
        total=Sum('cost_price', output_field=DecimalField(max_digits=15, decimal_places=2))
    )['total'] or Decimal('0.00')

    accessory_count = accessories.aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
    accessory_cost_value = accessories.aggregate(
        total=Sum(F('purchase_price') * F('quantity'), output_field=DecimalField(max_digits=15, decimal_places=2))
    )['total'] or Decimal('0.00')

    context = {
        'shop': shop,
        'phones': phones,
        'accessories': accessories,
        'phone_count': phones.count(),
        'accessory_count': accessory_count,
        'phone_cost_value': phone_cost_value,
        'accessory_cost_value': accessory_cost_value,
        'total_value': phone_cost_value + accessory_cost_value,
    }
    return render(request, 'inventory/shop_detail.html', context)


@login_required
def accessory_search(request):
    """Aksessuarlar AJAX qidirish"""
    code_query = request.GET.get('code', '').strip()
    name_query = request.GET.get('name', '').strip()
    shop_id = request.GET.get('shop_id', '')
    stock_level = request.GET.get('stock_level', '')
    page = request.GET.get('page', 1)

    accessories = Accessory.objects.select_related('supplier', 'created_by', 'shop')

    if code_query:
        accessories = accessories.filter(code__icontains=code_query)
    if name_query:
        accessories = accessories.filter(name__icontains=name_query)
    if shop_id:
        try:
            accessories = accessories.filter(shop_id=int(shop_id))
        except (ValueError, TypeError):
            pass

    stock_filters = {
        'low': Q(quantity__lt=5),
        'medium': Q(quantity__gte=5, quantity__lte=10),
        'high': Q(quantity__gt=10)
    }
    if stock_level in stock_filters:
        accessories = accessories.filter(stock_filters[stock_level])

    accessories = accessories.order_by('-created_at')

    total_accessory_cost = accessories.aggregate(
        total=Sum(F('purchase_price') * F('quantity'), output_field=DecimalField(max_digits=15, decimal_places=2))
    )['total'] or Decimal('0.00')

    paginator = Paginator(accessories, 12)

    try:
        accessories_page = paginator.page(page)
    except Exception:
        accessories_page = paginator.page(1)

    results = {
        'accessories': [{
            'id': acc.id,
            'name': acc.name,
            'code': acc.code,
            'image': acc.image.url if acc.image else None,
            'purchase_price': float(acc.purchase_price),
            'sale_price': float(acc.sale_price),
            'quantity': acc.quantity,
            'total_cost': float(acc.purchase_price * acc.quantity),
            'shop_name': acc.shop.name,
            'supplier_name': acc.supplier.name if acc.supplier else None,
            'created_at': acc.created_at.strftime('%d.%m.%Y') if acc.created_at else '',
            'can_edit': request.user == acc.shop.owner,
        } for acc in accessories_page],
        'has_next': accessories_page.has_next(),
        'has_previous': accessories_page.has_previous(),
        'current_page': accessories_page.number,
        'total_pages': paginator.num_pages,
        'total_count': paginator.count,
        'total_accessory_cost': float(total_accessory_cost),
    }

    return JsonResponse(results)


@login_required
def search(request):
    """Telefonlar AJAX qidirish"""
    imei_query = request.GET.get('imei', '').strip()
    status_filter = request.GET.get('status', '')
    shop_id = request.GET.get('shop_id', '')
    model_query = request.GET.get('model', '').strip()
    page = request.GET.get('page', 1)

    phones = Phone.objects.select_related('phone_model', 'memory_size', 'shop', 'created_by')

    if imei_query:
        phones = phones.filter(imei__icontains=imei_query)
    if status_filter:
        phones = phones.filter(status=status_filter)
    if shop_id:
        try:
            phones = phones.filter(shop_id=int(shop_id))
        except (ValueError, TypeError):
            pass
    if model_query:
        phones = phones.filter(phone_model__model_name__icontains=model_query)

    phones = phones.order_by('-created_at')
    paginator = Paginator(phones, 10)

    try:
        phones_page = paginator.page(page)
    except Exception:
        phones_page = paginator.page(1)

    results = {
        'phones': [{
            'id': phone.id,
            'phone_model': str(phone.phone_model),
            'memory_size': str(phone.memory_size),
            'imei': phone.imei or 'N/A',
            'status': phone.status,
            'status_display': phone.get_status_display(),
            'shop_name': phone.shop.name,
            'cost_price': float(phone.cost_price),
            'sale_price': float(phone.sale_price) if phone.sale_price else None,
            'condition_percentage': phone.condition_percentage,
            'created_at': phone.created_at.isoformat() if phone.created_at else '',
            'created_by': phone.created_by.username if phone.created_by else None,
            'image': phone.image.url if phone.image else None,
            'can_edit': request.user == phone.shop.owner,
        } for phone in phones_page],
        'has_next': phones_page.has_next(),
        'has_previous': phones_page.has_previous(),
        'current_page': phones_page.number,
        'total_pages': paginator.num_pages,
        'total_count': paginator.count,
    }

    return JsonResponse(results)


@require_GET
@login_required
def phone_details_api(request, phone_id):
    """Telefon tafsilotlari API"""
    try:
        phone = get_object_or_404(Phone, id=phone_id, shop__owner=request.user)
        data = {
            'success': True,
            'phone': {
                'model': str(phone.phone_model),
                'memory': str(phone.memory_size),
                'imei': phone.imei or 'N/A',
                'condition': phone.condition_percentage,
                'cost_price': str(phone.cost_price),
                'repair_cost': str(phone.repair_cost),
            }
        }
        return JsonResponse(data)
    except Phone.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Telefon topilmadi'}, status=404)


@login_required
@boss_or_finance_required
def phone_detail(request, pk):
    """Telefon tafsilotlari - TO'LIQ YANGILANGAN"""
    phone = get_object_or_404(Phone, pk=pk)

    # ========== INIT ==========
    phone_sale = None
    phone_return = None
    phone_exchange_new = None
    phone_exchange_old = None
    customer_debts = []
    actual_sale_info = None  # ✅ ASOSIY SOTISH MA'LUMOTLARI

    try:
        from sales.models import PhoneSale, PhoneReturn, PhoneExchange, Debt

        # ========== 1️⃣ TELEFON SOTISH ORQALI ==========
        phone_sale = PhoneSale.objects.select_related(
            'customer', 'salesman'
        ).filter(phone=phone).first()

        if phone_sale:
            actual_sale_info = {
                'type': 'phone_sale',
                'sale_price': phone_sale.sale_price,
                'salesman': phone_sale.salesman,
                'customer': phone_sale.customer,
                'sale_date': phone_sale.sale_date,
                'cash_amount': phone_sale.cash_amount,
                'card_amount': phone_sale.card_amount,
                'credit_amount': phone_sale.credit_amount,
                'debt_amount': phone_sale.debt_amount,
                'is_returned': phone_sale.is_returned,
            }

            # Qaytarish
            phone_return = PhoneReturn.objects.select_related(
                'created_by'
            ).filter(phone_sale=phone_sale).first()

            # Qarzlar
            if phone_sale.debt_amount > 0:
                customer_debts = Debt.objects.filter(
                    customer=phone_sale.customer,
                    status='active'
                ).select_related('creditor').prefetch_related('payments')

        # ========== 2️⃣ ALMASHTIRISH ORQALI (Yangi telefon sifatida) ==========
        phone_exchange_new = PhoneExchange.objects.select_related(
            'old_phone_model', 'old_phone_memory', 'salesman', 'customer'
        ).filter(new_phone=phone).first()

        if phone_exchange_new:
            # ✅ Agar telefon ALMASHTIRISH orqali sotilgan bo'lsa
            actual_sale_info = {
                'type': 'phone_exchange',
                'sale_price': phone_exchange_new.new_phone_price,
                'salesman': phone_exchange_new.salesman,
                'customer': phone_exchange_new.customer,
                'customer_name': phone_exchange_new.customer_name,
                'customer_phone': phone_exchange_new.customer_phone_number,
                'sale_date': phone_exchange_new.exchange_date,
                'cash_amount': phone_exchange_new.cash_amount,
                'card_amount': phone_exchange_new.card_amount,
                'credit_amount': phone_exchange_new.credit_amount,
                'debt_amount': phone_exchange_new.debt_amount,
                'exchange_type': phone_exchange_new.exchange_type,
                'old_phone_model': f"{phone_exchange_new.old_phone_model} {phone_exchange_new.old_phone_memory}",
                'old_phone_accepted_price': phone_exchange_new.old_phone_accepted_price,
                'price_difference': phone_exchange_new.price_difference,
            }

            # Qarzlar (almashtirish orqali)
            if phone_exchange_new.customer and phone_exchange_new.debt_amount > 0:
                customer_debts = Debt.objects.filter(
                    customer=phone_exchange_new.customer,
                    status='active'
                ).select_related('creditor').prefetch_related('payments')

        # ========== 3️⃣ ALMASHTIRISH ORQALI (Eski telefon sifatida qabul qilingan) ==========
        phone_exchange_old = PhoneExchange.objects.select_related(
            'new_phone__phone_model', 'new_phone__memory_size',
            'salesman', 'customer'
        ).filter(created_old_phone=phone).first()

    except ImportError:
        pass

    # ========== TIMELINE ==========
    timeline = [{
        'date': phone.created_at,
        'event': 'Telefon qo\'shildi',
        'description': f'{phone.get_source_type_display()} orqali qo\'shildi',
        'user': phone.created_by,
        'type': 'created'
    }]

    # Timeline: Oddiy sotish
    if phone_sale:
        timeline.append({
            'date': phone_sale.sale_date,
            'event': 'Sotildi',
            'description': f'{phone_sale.customer.name} ga ${phone_sale.sale_price} ga sotildi',
            'user': phone_sale.salesman,
            'type': 'sold'
        })

    # Timeline: Almashtirish (yangi telefon)
    if phone_exchange_new:
        timeline.append({
            'date': phone_exchange_new.exchange_date,
            'event': 'Almashtirish orqali sotildi',
            'description': f'{phone_exchange_new.customer_name} ga ${phone_exchange_new.new_phone_price} ga sotildi. Eski telefon: {phone_exchange_new.old_phone_model} {phone_exchange_new.old_phone_memory}',
            'user': phone_exchange_new.salesman,
            'type': 'exchanged_new'
        })

    # Timeline: Almashtirish (eski telefon)
    if phone_exchange_old:
        timeline.append({
            'date': phone_exchange_old.exchange_date,
            'event': 'Almashtirish orqali qabul qilindi',
            'description': f'{phone_exchange_old.customer_name} dan ${phone_exchange_old.old_phone_accepted_price} ga qabul qilindi. Yangi telefon: {phone_exchange_old.new_phone}',
            'user': phone_exchange_old.salesman,
            'type': 'exchanged_old'
        })

    # Timeline: Qaytarish
    if phone_return:
        timeline.append({
            'date': phone_return.return_date,
            'event': 'Qaytarildi',
            'description': f'${phone_return.return_amount} qaytarildi. Sabab: {phone_return.reason[:50]}',
            'user': phone_return.created_by,
            'type': 'returned'
        })

    # Timeline sortlash
    timeline.sort(key=lambda x: x['date'] if x['date'] else phone.created_at)

    # ========== MOLIYAVIY XULOSA ==========
    financial_summary = {
        'total_cost': phone.cost_price,
        'purchase_price': phone.purchase_price,
        'imei_cost': phone.imei_cost,
        'repair_cost': phone.repair_cost,
        'sale_price': Decimal('0.00'),
        'profit': Decimal('0.00'),
        'return_amount': Decimal('0.00'),
        'final_profit': Decimal('0.00')
    }

    # ✅ Asosiy sotish narxini aniqlash
    if actual_sale_info:
        financial_summary['sale_price'] = actual_sale_info['sale_price']
        financial_summary['profit'] = financial_summary['sale_price'] - financial_summary['total_cost']

        # Qaytarish summasi
        if phone_return:
            financial_summary['return_amount'] = phone_return.return_amount

        financial_summary['final_profit'] = financial_summary['profit'] - financial_summary['return_amount']

    # ========== CONTEXT ==========
    context = {
        'phone': phone,
        'phone_sale': phone_sale,
        'phone_return': phone_return,
        'phone_exchange_new': phone_exchange_new,
        'phone_exchange_old': phone_exchange_old,
        'customer_debts': customer_debts,
        'timeline': timeline,
        'financial_summary': financial_summary,
        'actual_sale_info': actual_sale_info,  # ✅ ASOSIY SOTISH MA'LUMOTLARI
        'can_edit': request.user == phone.shop.owner,
    }

    return render(request, 'inventory/phone_detail.html', context)


@login_required
def search_external_seller_api(request):
    """Tashqi sotuvchilarni qidirish API"""
    query = request.GET.get("q", "").strip()

    if len(query) < 2:
        return JsonResponse([], safe=False)

    sellers = ExternalSeller.objects.filter(
        Q(phone_number__icontains=query) | Q(name__icontains=query)
    ).order_by('name')[:10]

    data = [{
        "id": seller.id,
        "name": seller.name,
        "phone_number": seller.phone_number,
        "display_text": f"{seller.name} - {seller.phone_number}",
        "created_at": seller.created_at.strftime('%d.%m.%Y') if seller.created_at else ""
    } for seller in sellers]

    return JsonResponse(data, safe=False)


@login_required
def check_external_seller_phone_api(request):
    """Tashqi sotuvchi telefon raqami tekshirish API"""
    phone = request.GET.get("phone", "").strip()

    if len(phone) < 7:
        return JsonResponse({'exists': False, 'message': ''})

    try:
        seller = ExternalSeller.objects.get(phone_number=phone)
        return JsonResponse({
            'exists': True,
            'seller': {
                'id': seller.id,
                'name': seller.name,
                'phone_number': seller.phone_number,
                'created_at': seller.created_at.strftime('%d.%m.%Y') if seller.created_at else ''
            },
            'message': f'Bu telefon raqam allaqachon mavjud: {seller.name}'
        })
    except ExternalSeller.DoesNotExist:
        return JsonResponse({
            'exists': False,
            'message': 'Telefon raqam mavjud emas. Yangi tashqi sotuvchi yaratiladi.'
        })


@login_required
def check_daily_seller_phone_api(request):
    """Kunlik sotuvchi telefon raqami tekshirish API"""
    phone = request.GET.get("phone", "").strip()

    if len(phone) < 7:
        return JsonResponse({'exists': False, 'message': ''})

    try:
        seller = DailySeller.objects.get(phone_number=phone)
        return JsonResponse({
            'exists': True,
            'seller': {
                'id': seller.id,
                'name': seller.name,
                'phone_number': seller.phone_number,
                'created_at': seller.created_at.strftime('%d.%m.%Y') if seller.created_at else ''
            },
            'message': f'Bu telefon raqam allaqachon mavjud: {seller.name}'
        })
    except DailySeller.DoesNotExist:
        return JsonResponse({
            'exists': False,
            'message': 'Telefon raqam mavjud emas. Yangi kunlik sotuvchi yaratiladi.'
        })

@login_required
def supplier_list(request):
    """Taminotchilar ro'yxati"""
    suppliers = Supplier.objects.all().order_by('-created_at')

    search_query = request.GET.get('search', '').strip()
    if search_query:
        suppliers = suppliers.filter(
            Q(name__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(notes__icontains=search_query)
        )

    paginator = Paginator(suppliers, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # ✅ TO'G'RI STATISTIKA - Barcha supplierlar bo'yicha
    all_suppliers = Supplier.objects.all()

    suppliers_with_debt = 0
    total_debt = Decimal('0.00')
    total_paid = Decimal('0.00')

    for supplier in all_suppliers:
        balance = supplier.balance
        if balance > 0:
            suppliers_with_debt += 1
        total_debt += (supplier.initial_debt + supplier.total_debt)  # ✅ To'g'ri hisoblash
        total_paid += supplier.total_paid

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'total_suppliers': all_suppliers.count(),
        'suppliers_with_debt': suppliers_with_debt,
        'total_debt': total_debt,
        'total_paid': total_paid,
    }

    return render(request, 'inventory/supplier_list.html', context)


@login_required
def supplier_detail(request, pk):
    """Taminotchi tafsilotlari - qarzlar va to'lovlar"""
    supplier = get_object_or_404(Supplier, pk=pk)

    # Qarzda turgan telefonlar
    debt_phones = supplier.get_debt_phones()

    # To'langan telefonlar (pagination)
    paid_phones_all = supplier.get_paid_phones()

    # Pagination - har sahifada 20 ta
    paginator = Paginator(paid_phones_all, 20)
    page_number = request.GET.get('page')
    paid_phones_page = paginator.get_page(page_number)

    # To'lov tarixi
    payments = SupplierPayment.objects.filter(
        supplier=supplier
    ).select_related('created_by').prefetch_related('details__phone').order_by('-payment_date')

    # Statistika
    total_phones = Phone.objects.filter(
        supplier=supplier,
        source_type='supplier'
    ).count()

    # Balansni aniq hisoblash
    balance = supplier.balance

    context = {
        'supplier': supplier,
        'debt_phones': debt_phones,
        'paid_phones': paid_phones_page,
        'paid_phones_total': paid_phones_all.count(),
        'payments': payments,
        'total_phones': total_phones,
        'balance': balance,
        'can_edit': can_edit_inventory(request.user),
    }

    return render(request, 'inventory/supplier_detail.html', context)


@login_required
@boss_or_finance_required
def supplier_payment_create(request, supplier_id):
    """Taminotchiga to'lov qilish"""
    supplier = get_object_or_404(Supplier, pk=supplier_id)

    # Taminotchining balansini tekshirish
    if supplier.balance <= 0:
        messages.warning(request, "Bu taminotchida qarz yo'q!")
        return redirect('inventory:supplier_detail', pk=supplier_id)

    if request.method == 'POST':
        form = SupplierPaymentForm(request.POST, supplier=supplier)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # 1. To'lovni saqlash
                    payment = form.save(commit=False)
                    payment.supplier = supplier
                    payment.created_by = request.user
                    payment.save()

                    amount_remaining = payment.amount
                    initial_payment_made = Decimal('0.00')
                    paid_phones_count = 0

                    # 2. Telefonlarni tanlash
                    if payment.payment_type == 'general':
                        phones = supplier.get_debt_phones()
                    else:
                        phones = form.cleaned_data.get('selected_phones', [])
                        phones = sorted(phones, key=lambda p: p.created_at) if phones else []

                    # 3. Telefonlar qarzini to'lash
                    for phone in phones:
                        if amount_remaining <= 0:
                            break

                        # Telefonning qolgan qarzini hisoblash
                        remaining_debt = phone.debt_balance
                        if remaining_debt <= 0:
                            continue

                        # To'lanadigan summani aniqlash
                        payment_for_phone = min(amount_remaining, remaining_debt)

                        # Telefonning to'langan summasini yangilash
                        phone.paid_amount += payment_for_phone
                        phone.debt_balance = phone.cost_price - phone.paid_amount
                        phone.payment_status = 'paid' if phone.debt_balance <= 0 else 'partial'
                        phone.save(update_fields=['paid_amount', 'debt_balance', 'payment_status'])

                        # To'lov tafsilotini saqlash
                        SupplierPaymentDetail.objects.create(
                            payment=payment,
                            phone=phone,
                            amount=payment_for_phone,
                            previous_balance=remaining_debt,
                            new_balance=phone.debt_balance
                        )

                        amount_remaining -= payment_for_phone
                        paid_phones_count += 1

                    # 4. Boshlang'ich qarzni to'lash
                    if amount_remaining > 0 and supplier.initial_debt > 0:
                        initial_payment_made = min(amount_remaining, supplier.initial_debt)
                        supplier.initial_debt -= initial_payment_made
                        amount_remaining -= initial_payment_made

                    # 5. Taminotchining umumiy to'langan summasini yangilash
                    supplier.total_paid += (payment.amount - amount_remaining)
                    supplier.save(update_fields=['total_paid', 'initial_debt'])

                    # 6. Taminotchining umumiy qarzini yangilash
                    supplier.update_total_debt()
                    supplier.refresh_from_db()

                    # 7. Muvaffaqiyat xabari
                    success_msg = f"{supplier.name} ga ${payment.amount} to'lov qilindi!"
                    if paid_phones_count > 0:
                        success_msg += f" {paid_phones_count} ta telefon qarzi to'landi."
                    if initial_payment_made > 0:
                        success_msg += f" Boshlang'ich qarzdan ${initial_payment_made} to'landi."
                    if amount_remaining > 0:
                        success_msg += f" ${amount_remaining} keyingi to'lovga o'tkazildi."
                    success_msg += f" Qoldiq qarz: ${supplier.balance}"

                    messages.success(request, success_msg)
                    return redirect('inventory:supplier_detail', pk=supplier_id)

            except Exception as e:
                messages.error(request, f"To'lov qilishda xatolik: {str(e)}")
        else:
            for field, errors in form.errors.items():
                field_label = form.fields[field].label if field in form.fields else field.replace('_', ' ').title()
                for error in errors:
                    messages.error(request, f"{field_label}: {error}")
    else:
        form = SupplierPaymentForm(supplier=supplier)

    context = {
        'form': form,
        'supplier': supplier,
        'debt_phones': supplier.get_debt_phones(),
        'is_edit': False,
    }

    return render(request, 'inventory/supplier_payment_form.html', context)


@login_required
def supplier_payment_detail(request, payment_id):
    """To'lov tafsilotlari"""
    payment = get_object_or_404(
        SupplierPayment.objects.prefetch_related('details__phone'),
        pk=payment_id
    )

    context = {
        'payment': payment,
    }

    return render(request, 'inventory/supplier_payment_detail.html', context)


@login_required
@boss_or_finance_required
def supplier_create(request):
    """Yangi taminotchi qo'shish"""
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            supplier = form.save()

            success_msg = f"{supplier.name} muvaffaqiyatli qo'shildi!"
            if supplier.initial_debt > 0:
                success_msg += f" Boshlang'ich qarz: ${supplier.initial_debt}"

            messages.success(request, success_msg)
            return redirect('inventory:supplier_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, f"{error}")
                    else:
                        field_label = form.fields.get(field, {}).label or field
                        messages.error(request, f"{field_label}: {error}")
    else:
        form = SupplierForm()

    context = {
        'form': form,
        'title': 'Yangi Taminotchi',
        'is_edit': False
    }
    return render(request, 'inventory/supplier_form.html', context)


@login_required
@boss_or_finance_required
def supplier_update(request, pk):
    """Taminotchini tahrirlash"""
    supplier = get_object_or_404(Supplier, pk=pk)

    # Oldingi qarzni saqlash
    old_initial_debt = supplier.initial_debt

    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            supplier = form.save()

            success_msg = f"{supplier.name} ma'lumotlari yangilandi!"

            # Agar boshlang'ich qarz o'zgargan bo'lsa
            if old_initial_debt != supplier.initial_debt:
                debt_change = supplier.initial_debt - old_initial_debt
                if debt_change > 0:
                    success_msg += f" Boshlang'ich qarz +${debt_change} oshdi."
                elif debt_change < 0:
                    success_msg += f" Boshlang'ich qarz ${abs(debt_change)} kamaydi."

            messages.success(request, success_msg)
            return redirect('inventory:supplier_detail', pk=pk)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, f"{error}")
                    else:
                        field_label = form.fields.get(field, {}).label or field
                        messages.error(request, f"{field_label}: {error}")
    else:
        form = SupplierForm(instance=supplier)

    context = {
        'form': form,
        'supplier': supplier,
        'title': 'Taminotchini Tahrirlash',
        'is_edit': True
    }
    return render(request, 'inventory/supplier_form.html', context)


@login_required
@boss_or_finance_required
def supplier_delete(request, pk):
    """Taminotchini o'chirish"""
    supplier = get_object_or_404(Supplier, pk=pk)

    # Qarzli telefonlar borligini tekshirish
    if supplier.balance > 0:
        messages.error(
            request,
            f"Bu taminotchida ${supplier.balance} qarz bor. "
            "Avval qarzni to'lang!"
        )
        return redirect('inventory:supplier_detail', pk=pk)

    # Bog'liq telefonlar borligini tekshirish
    phones_count = Phone.objects.filter(supplier=supplier).count()
    if phones_count > 0:
        messages.error(
            request,
            f"Bu taminotchiga {phones_count} ta telefon bog'langan. "
            "Avval telefonlarni o'chiring yoki boshqa taminotchiga o'tkazing!"
        )
        return redirect('inventory:supplier_detail', pk=pk)

    if request.method == 'POST':
        supplier_name = supplier.name
        supplier.delete()
        messages.success(request, f"{supplier_name} muvaffaqiyatli o'chirildi!")
        return redirect('inventory:supplier_list')

    context = {
        'supplier': supplier
    }
    return render(request, 'inventory/supplier_confirm_delete.html', context)


@login_required
@boss_or_finance_required
def supplier_payment_update(request, payment_id):
    """To'lovni tahrirlash"""
    payment = get_object_or_404(SupplierPayment, pk=payment_id)
    supplier = payment.supplier

    if request.method == 'POST':
        # Eski to'lov ma'lumotlarini saqlash
        old_amount = payment.amount

        form = SupplierPaymentForm(request.POST, instance=payment, supplier=supplier)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # 1. ESKI TO'LOVNI BEKOR QILISH

                    # Telefon to'lovlarini qaytarish
                    for detail in payment.details.all():
                        phone = detail.phone
                        phone.paid_amount -= detail.amount
                        phone.save(update_fields=['paid_amount'])

                    # Boshlang'ich qarzga to'langan pulni qaytarish
                    phone_payments_total = payment.details.aggregate(
                        total=Sum('amount')
                    )['total'] or Decimal('0.00')

                    initial_paid = old_amount - phone_payments_total
                    if initial_paid > 0:
                        supplier.initial_debt += initial_paid

                    # Supplier total_paid ni kamaytirish
                    supplier.total_paid -= old_amount
                    supplier.save(update_fields=['total_paid', 'initial_debt'])

                    # 2. ESKI TAFSILOTLARNI O'CHIRISH
                    payment.details.all().delete()

                    # 3. YANGI TO'LOVNI QO'LLASH
                    new_payment = form.save(commit=False)
                    new_payment.save()

                    amount_remaining = new_payment.amount
                    initial_payment_made = Decimal('0.00')
                    paid_phones_count = 0

                    # Yangi to'lovni telefonlarga tarqatish
                    if new_payment.payment_type == 'general':
                        phones = supplier.get_debt_phones()
                    else:
                        phones = form.cleaned_data.get('selected_phones', [])
                        phones = sorted(phones, key=lambda p: p.created_at) if phones else []

                    for phone in phones:
                        if amount_remaining <= 0:
                            break

                        remaining_debt = phone.cost_price - phone.paid_amount
                        if remaining_debt <= 0:
                            continue

                        payment_for_phone = min(amount_remaining, remaining_debt)

                        phone.paid_amount += payment_for_phone
                        phone.save(update_fields=['paid_amount'])

                        SupplierPaymentDetail.objects.create(
                            payment=new_payment,
                            phone=phone,
                            amount=payment_for_phone,
                            previous_balance=remaining_debt + payment_for_phone,
                            new_balance=remaining_debt
                        )

                        amount_remaining -= payment_for_phone
                        paid_phones_count += 1

                    # Boshlang'ich qarzga to'lov
                    if amount_remaining > 0 and supplier.initial_debt > 0:
                        initial_payment_made = min(amount_remaining, supplier.initial_debt)
                        supplier.initial_debt -= initial_payment_made
                        amount_remaining -= initial_payment_made

                    # Supplier yangilash
                    supplier.total_paid += (new_payment.amount - amount_remaining)
                    supplier.save(update_fields=['total_paid', 'initial_debt'])
                    supplier.update_total_debt()

                    # Xabar
                    success_msg = f"To'lov muvaffaqiyatli yangilandi!"
                    if paid_phones_count > 0:
                        success_msg += f" {paid_phones_count} ta telefon qarzi to'landi."
                    if initial_payment_made > 0:
                        success_msg += f" Boshlang'ich qarzdan ${initial_payment_made} to'landi."
                    if amount_remaining > 0:
                        success_msg += f" ${amount_remaining} overpayment keyingi to'lovga o'tkazildi."

                    messages.success(request, success_msg)
                    return redirect('inventory:supplier_detail', pk=supplier.id)

            except Exception as e:
                messages.error(request, f"Yangilashda xatolik: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{error}")
    else:
        form = SupplierPaymentForm(instance=payment, supplier=supplier)

    context = {
        'form': form,
        'payment': payment,
        'supplier': supplier,
        'debt_phones': supplier.get_debt_phones(),
        'is_edit': True,
    }

    return render(request, 'inventory/supplier_payment_form.html', context)


@login_required
@boss_or_finance_required
def supplier_payment_delete(request, payment_id):
    """To'lovni o'chirish"""
    payment = get_object_or_404(SupplierPayment, pk=payment_id)
    supplier = payment.supplier

    if request.method == 'POST':
        try:
            with transaction.atomic():
                # 1. Telefon to'lovlarini qaytarish
                total_phone_payment = Decimal('0.00')
                for detail in payment.details.all():
                    phone = detail.phone
                    phone.paid_amount -= detail.amount
                    phone.save(update_fields=['paid_amount'])
                    total_phone_payment += detail.amount

                # 2. Boshlang'ich qarzga to'langan pulni qaytarish
                initial_paid = payment.amount - total_phone_payment
                if initial_paid > 0:
                    supplier.initial_debt += initial_paid

                # 3. Supplier total_paid ni kamaytirish
                supplier.total_paid -= payment.amount
                supplier.save(update_fields=['total_paid', 'initial_debt'])

                # 4. To'lovni o'chirish
                payment_amount = payment.amount
                payment.delete()

                # 5. Supplier qarzini yangilash
                supplier.update_total_debt()
                supplier.refresh_from_db()

                messages.success(
                    request,
                    f"${payment_amount} to'lov o'chirildi. Qoldiq qarz: ${supplier.balance}"
                )
                return redirect('inventory:supplier_detail', pk=supplier.id)

        except Exception as e:
            messages.error(request, f"O'chirishda xatolik: {str(e)}")

    context = {
        'payment': payment,
        'supplier': supplier,
    }

    return render(request, 'inventory/supplier_payment_confirm_delete.html', context)


@login_required
def check_imei_api(request):
    """IMEI tekshirish API"""
    imei = request.GET.get('imei', '').strip()
    current_id = request.GET.get('current_id', '').strip()

    if not imei or len(imei) != 15:
        return JsonResponse({'exists': False})

    # IMEI ni tekshirish
    phones = Phone.objects.filter(imei=imei)

    # Agar tahrirlash bo'lsa - o'sha telefonni chiqarib tashlash
    if current_id and current_id.isdigit():
        phones = phones.exclude(id=int(current_id))

    if phones.exists():
        phone = phones.select_related('phone_model', 'memory_size', 'shop').first()
        return JsonResponse({
            'exists': True,
            'phone': {
                'id': phone.id,
                'phone_model': phone.phone_model.model_name,
                'memory_size': str(phone.memory_size),
                'imei': phone.imei,
                'shop_name': phone.shop.name,
                'status': phone.status,
                'status_display': phone.get_status_display(),
                'cost_price': str(phone.cost_price) if phone.cost_price else None,
            }
        })

    return JsonResponse({'exists': False})


# views.py ga qo'shish kerak bo'lgan funksiyalar




@login_required
def export_phones_to_excel(request):
    """Telefonlarni to'liq ma'lumotlar bilan Excel ga eksport qilish"""

    # Filtrlar
    status_filter = request.GET.get('status', '').strip()
    shop_id = request.GET.get('shop_id', '').strip()
    model_query = request.GET.get('model', '').strip()
    imei_query = request.GET.get('imei', '').strip()

    # Telefonlarni olish
    phones = Phone.objects.select_related(
        'phone_model', 'memory_size', 'shop', 'supplier',
        'external_seller', 'daily_seller', 'created_by'
    ).order_by('-created_at')

    # Filtrlarni qo'llash
    if status_filter:
        phones = phones.filter(status=status_filter)
    if shop_id:
        phones = phones.filter(shop_id=shop_id)
    if model_query:
        phones = phones.filter(phone_model__id=model_query)
    if imei_query:
        phones = phones.filter(imei__icontains=imei_query)

    # Excel fayl yaratish
    wb = Workbook()
    ws = wb.active
    ws.title = "Telefonlar"

    # Sarlavha stillari
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Chegaralar
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sarlavhalar
    headers = [
        '№',
        'Do\'kon',
        'Telefon Modeli',
        'Xotira',
        'IMEI',
        'Holati (%)',
        'Status',
        'Sotib olingan ($)',
        'IMEI xarajat ($)',
        'Ta\'mirlash ($)',
        'Tan narx ($)',
        'Sotish narxi ($)',
        'Manba',
        'Taminotchi/Sotuvchi',
        'To\'lov holati',
        'To\'langan ($)',
        'Qarz ($)',
        'Qo\'shilgan sana',
        'Qo\'shgan',
        'Izoh'
    ]

    # Sarlavhalarni yozish
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Ma'lumotlarni yozish
    row_num = 2
    for idx, phone in enumerate(phones, 1):
        # Manba aniqlash
        if phone.source_type == 'supplier':
            source_name = phone.supplier.name if phone.supplier else 'N/A'
        elif phone.source_type == 'external_seller':
            source_name = phone.external_seller.name if phone.external_seller else 'N/A'
        elif phone.source_type == 'daily_seller':
            source_name = phone.daily_seller.name if phone.daily_seller else 'N/A'
        elif phone.source_type == 'exchange':
            source_name = phone.original_owner_name or 'N/A'
        else:
            source_name = 'N/A'

        # Qatorni to'ldirish
        row_data = [
            idx,
            phone.shop.name,
            phone.phone_model.model_name,
            str(phone.memory_size),
            phone.imei or 'N/A',
            phone.condition_percentage,
            phone.get_status_display(),
            float(phone.purchase_price),
            float(phone.imei_cost),
            float(phone.repair_cost),
            float(phone.cost_price),
            float(phone.sale_price) if phone.sale_price else '',
            phone.get_source_type_display(),
            source_name,
            phone.get_payment_status_display() if phone.source_type == 'supplier' else 'To\'langan',
            float(phone.paid_amount) if phone.source_type == 'supplier' else float(phone.cost_price),
            float(phone.debt_balance) if phone.source_type == 'supplier' else 0,
            phone.created_at.strftime('%d.%m.%Y') if phone.created_at else '',
            phone.created_by.username if phone.created_by else '',
            phone.note or ''
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Raqamlar uchun format
            if col_num in [8, 9, 10, 11, 12, 16, 17]:  # Dollar ustunlari
                if value and value != '':
                    cell.number_format = '$#,##0.00'

        row_num += 1

    # Ustun kengliklarini sozlash
    column_widths = {
        'A': 5,  # №
        'B': 15,  # Do'kon
        'C': 18,  # Model
        'D': 10,  # Xotira
        'E': 18,  # IMEI
        'F': 10,  # Holat
        'G': 15,  # Status
        'H': 12,  # Sotib olingan
        'I': 12,  # IMEI xarajat
        'J': 12,  # Ta'mirlash
        'K': 12,  # Tan narx
        'L': 12,  # Sotish narxi
        'M': 18,  # Manba
        'N': 20,  # Taminotchi
        'O': 15,  # To'lov holati
        'P': 12,  # To'langan
        'Q': 12,  # Qarz
        'R': 15,  # Sana
        'S': 15,  # Qo'shgan
        'T': 30  # Izoh
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Qatorlarning balandligini sozlash
    ws.row_dimensions[1].height = 30

    # Statistika qo'shish (pastga)
    stats_row = row_num + 2

    # Statistika sarlavhasi
    ws.cell(row=stats_row, column=1).value = "STATISTIKA"
    ws.cell(row=stats_row, column=1).font = Font(bold=True, size=12)
    stats_row += 1

    # Statistika ma'lumotlari
    total_phones = phones.count()
    total_cost = sum(float(p.cost_price) for p in phones)
    total_debt = sum(float(p.debt_balance) for p in phones if p.source_type == 'supplier')
    total_paid = sum(float(p.paid_amount) for p in phones if p.source_type == 'supplier')

    stats_data = [
        ('Jami telefonlar:', total_phones),
        ('Jami tan narx:', f'${total_cost:,.2f}'),
        ('Jami qarz:', f'${total_debt:,.2f}'),
        ('Jami to\'langan:', f'${total_paid:,.2f}'),
    ]

    for label, value in stats_data:
        ws.cell(row=stats_row, column=1).value = label
        ws.cell(row=stats_row, column=1).font = Font(bold=True)
        ws.cell(row=stats_row, column=2).value = value
        stats_row += 1

    # HTTP javob
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # Fayl nomi - filtrlar bilan
    filename = 'telefonlar'
    if status_filter:
        filename += f'_{status_filter}'
    if shop_id:
        try:
            shop = Shop.objects.get(id=shop_id)
            filename += f'_{shop.name}'
        except:
            pass

    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'

    wb.save(response)
    return response


@login_required
def export_phones_simple(request):
    """Oddiy variantda faqat Model va IMEI"""

    # Filtrlar
    status_filter = request.GET.get('status', '').strip()
    shop_id = request.GET.get('shop_id', '').strip()
    model_query = request.GET.get('model', '').strip()
    imei_query = request.GET.get('imei', '').strip()

    phones = Phone.objects.select_related('phone_model', 'memory_size', 'shop').order_by('-created_at')

    if status_filter:
        phones = phones.filter(status=status_filter)
    if shop_id:
        phones = phones.filter(shop_id=shop_id)
    if model_query:
        phones = phones.filter(phone_model__id=model_query)
    if imei_query:
        phones = phones.filter(imei__icontains=imei_query)

    # Excel yaratish
    wb = Workbook()
    ws = wb.active
    ws.title = "Telefonlar"

    # Sarlavhalar
    headers = ['№', 'Telefon Modeli', 'Xotira', 'IMEI', 'Do\'kon', 'Status']

    # Sarlavha stili
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Ma'lumotlar
    for idx, phone in enumerate(phones, 1):
        row_data = [
            idx,
            phone.phone_model.model_name,
            str(phone.memory_size),
            phone.imei or 'N/A',
            phone.shop.name,
            phone.get_status_display()
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=idx + 1, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Ustun kengliklari
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    # Qator balandligi
    ws.row_dimensions[1].height = 25

    # HTTP javob
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # Fayl nomi
    filename = 'telefonlar_oddiy'
    if status_filter:
        filename += f'_{status_filter}'

    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'

    wb.save(response)
    return response